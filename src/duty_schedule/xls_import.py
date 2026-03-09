from __future__ import annotations

from io import BytesIO
from typing import TYPE_CHECKING

from openpyxl import load_workbook

from duty_schedule.export.xls import SHIFT_LABELS
from duty_schedule.models import CarryOverState, ShiftType

if TYPE_CHECKING:
    from openpyxl.workbook import Workbook
    from openpyxl.worksheet.worksheet import Worksheet

LABEL_TO_SHIFT: dict[str, ShiftType] = {
    label.strip().lower(): ShiftType(key) for key, label in SHIFT_LABELS.items()
}

SCHED_SHEET_NAME = "График дежурств"

WORKING_SHIFTS = frozenset(
    {ShiftType.MORNING, ShiftType.EVENING, ShiftType.NIGHT, ShiftType.WORKDAY}
)


class XlsImportError(Exception):
    pass


def _find_sheet(wb: Workbook) -> Worksheet:
    for name in wb.sheetnames:
        if name == SCHED_SHEET_NAME:
            return wb[name]
    return wb.worksheets[0]


def _resolve_shift(raw: object) -> ShiftType:
    if raw is None:
        return ShiftType.DAY_OFF
    text = str(raw).strip().lower()
    if not text:
        return ShiftType.DAY_OFF
    return LABEL_TO_SHIFT.get(text, ShiftType.DAY_OFF)


def _find_day_columns(ws: Worksheet) -> tuple[int, int]:
    start_col = 3
    end_col = start_col
    for col in range(start_col, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val is not None and "итого" in str(val).strip().lower():
            break
        end_col = col
    if end_col < start_col:
        raise XlsImportError(
            "Не удалось определить диапазон дней в строке 2. "
            "Убедитесь, что файл содержит корректный график дежурств."
        )
    return start_col, end_col


def _build_carry_over(name: str, shifts: list[ShiftType]) -> CarryOverState:
    last_shift: ShiftType | None = None
    consecutive_working = 0
    consecutive_off = 0
    consecutive_same_shift = 0

    for s in reversed(shifts):
        if s in WORKING_SHIFTS:
            if consecutive_off > 0:
                break
            if last_shift is None:
                last_shift = s
                consecutive_same_shift = 1
            elif s == last_shift:
                consecutive_same_shift += 1
            else:
                break
            consecutive_working += 1
        else:
            if consecutive_working > 0:
                break
            consecutive_off += 1

    return CarryOverState(
        employee_name=name,
        last_shift=last_shift,
        consecutive_working=consecutive_working,
        consecutive_off=consecutive_off,
        consecutive_same_shift=consecutive_same_shift,
    )


def parse_carry_over_from_xls(file_bytes: bytes) -> list[CarryOverState]:
    try:
        wb = load_workbook(BytesIO(file_bytes), data_only=True)
    except Exception as exc:
        raise XlsImportError(f"Не удалось открыть файл: {exc}") from exc

    ws = _find_sheet(wb)

    try:
        start_col, end_col = _find_day_columns(ws)
    except XlsImportError:
        raise
    except Exception as exc:
        raise XlsImportError(f"Ошибка при разборе структуры файла: {exc}") from exc

    results: list[CarryOverState] = []
    row = 3
    while True:
        name_val = ws.cell(row=row, column=1).value
        if name_val is None or str(name_val).strip() == "":
            break
        name = str(name_val).strip()
        if name.startswith("'"):
            name = name[1:]

        shifts: list[ShiftType] = []
        for col in range(start_col, end_col + 1):
            cell_val = ws.cell(row=row, column=col).value
            shifts.append(_resolve_shift(cell_val))

        results.append(_build_carry_over(name, shifts))
        row += 1

    return results
