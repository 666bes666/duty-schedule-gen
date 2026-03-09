"""Экспорт расписания в формат Excel (.xlsx)."""

from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.worksheet import Worksheet

from duty_schedule.constants import (
    MONTHS_RU,
    SHIFT_COLORS_CELL,
    SHIFT_COLORS_HEADER,
)
from duty_schedule.models import City, DaySchedule, Employee, Schedule, ScheduleType
from duty_schedule.stats import (
    HOURS_NORMAL,
    HOURS_SHORT,
    EmployeeStats,
    build_assignments,
    compute_stats,
)

COLORS = {
    **SHIFT_COLORS_HEADER,
    "header": "404040",
    "name": "D9D9D9",
    "weekend": "F2F2F2",
    "ok": "E2EFDA",
    "warn": "FFF2CC",
    "bad": "FCE4D6",
    "over": "DDEBF7",
    "stat_header": "2F4F8F",
    "stat_section": "BDD7EE",
    "total_row": "595959",
}

CELL_COLORS = SHIFT_COLORS_CELL

WHITE_FONT_KEYS = {"evening", "header", "workday", "night", "vacation"}

THIN_SIDE = Side(style="thin", color="BFBFBF")
MEDIUM_SIDE = Side(style="medium", color="808080")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
MONDAY_BORDER = Border(left=MEDIUM_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SECTION_BORDER = Border(left=MEDIUM_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
SECTION_COLS = {3, 7, 11, 15}

SHIFT_LABELS = {
    "morning": "Утро",
    "evening": "Вечер",
    "night": "Ночь",
    "workday": "День",
    "day_off": "—",
    "vacation": "Отп",
}


def _sanitize_cell(value: str) -> str:
    if value and value[0] in ("=", "+", "-", "@"):
        return "'" + value
    return value


DAYS_RU = ["Пн", "Вт", "Ср", "Чт", "Пт", "Сб", "Вс"]
MONTHS_RU_SHORT = [
    "",
    "янв",
    "фев",
    "мар",
    "апр",
    "май",
    "июн",
    "июл",
    "авг",
    "сен",
    "окт",
    "ноя",
    "дек",
]

SCHED_SHEET = "График дежурств"
HELPER_SHEET = "_Данные"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")


def _font(bold: bool = False, white: bool = False, size: int = 10) -> Font:
    color = "FFFFFF" if white else "000000"
    return Font(bold=bold, color=color, name="Calibri", size=size)


def _align(horizontal: str = "center") -> Alignment:
    return Alignment(wrap_text=True, vertical="center", horizontal=horizontal)


def _darken(hex_color: str, factor: float = 0.88) -> str:
    r = int(int(hex_color[0:2], 16) * factor)
    g = int(int(hex_color[2:4], 16) * factor)
    b = int(int(hex_color[4:6], 16) * factor)
    return f"{r:02X}{g:02X}{b:02X}"


def _countif_working(sheet_name: str, day_range: str) -> str:
    s = f"'{sheet_name}'!"
    return (
        f'COUNTIF({s}{day_range},"Утро")'
        f'+COUNTIF({s}{day_range},"Вечер")'
        f'+COUNTIF({s}{day_range},"Ночь")'
        f'+COUNTIF({s}{day_range},"День")'
    )


def _is_working_array(sheet_name: str, day_range: str) -> str:
    s = f"'{sheet_name}'!"
    return (
        f'(({s}{day_range}="Утро")'
        f'+({s}{day_range}="Вечер")'
        f'+({s}{day_range}="Ночь")'
        f'+({s}{day_range}="День")>0)*1'
    )


def export_xls(schedule: Schedule, output_dir: Path, short_days: set[date] | None = None) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = output_dir / f"schedule_{schedule.config.year}_{schedule.config.month:02d}.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = SCHED_SHEET

    days = schedule.days
    employees = sorted(
        schedule.config.employees,
        key=lambda e: (
            0 if e.city == City.MOSCOW else 1,
            0 if not e.on_duty else 1,
            0 if e.schedule_type == ScheduleType.FIVE_TWO else 1,
            e.name,
        ),
    )
    assignments = build_assignments(schedule)
    production_days = schedule.metadata.get("production_working_days", 21)
    stats = compute_stats(schedule, assignments, production_days, employees, short_days)

    num_days = len(days)
    ld = get_column_letter(num_days + 2)

    _build_helper_sheet(wb, days, employees, production_days, short_days)
    _build_schedule_sheet(ws, days, employees, assignments, num_days, ld)

    ws_stat = wb.create_sheet(title="Статистика")
    _build_stats_sheet(ws_stat, stats, schedule, employees, num_days, ld)

    _add_legend(wb)

    wb.save(filename)
    return filename


def _build_helper_sheet(
    wb: Workbook,
    days: list[DaySchedule],
    employees: list[Employee],
    production_days: int,
    short_days: set[date] | None = None,
) -> None:
    ws = wb.create_sheet(title=HELPER_SHEET)
    ws.sheet_state = "hidden"

    _short = short_days or set()

    ws.cell(row=1, column=1, value="Дата")
    ws.cell(row=2, column=1, value="Часы")
    ws.cell(row=3, column=1, value="Выходной")
    ws.cell(row=4, column=1, value="Праздник")
    ws.cell(row=5, column=1, value="Произв.дней")
    ws.cell(row=5, column=2, value=production_days)

    for col_idx, day in enumerate(days, start=3):
        ws.cell(row=1, column=col_idx, value=day.date)
        ws.cell(row=2, column=col_idx, value=HOURS_SHORT if day.date in _short else HOURS_NORMAL)
        ws.cell(row=3, column=col_idx, value=1 if day.date.weekday() >= 5 else 0)
        ws.cell(row=4, column=col_idx, value=1 if day.is_holiday and day.date.weekday() < 5 else 0)

    for i, emp in enumerate(employees):
        row = 6 + i
        ws.cell(row=row, column=1, value=_sanitize_cell(emp.name))
        ws.cell(row=row, column=2, value=emp.workload_pct)


def _build_schedule_sheet(
    ws: Worksheet,
    days: list[DaySchedule],
    employees: list[Employee],
    assignments: dict[str, dict[date, str]],
    num_days: int,
    ld: str,
) -> None:
    total_col = num_days + 3
    hours_col = total_col + 1

    first_day = days[0].date
    month_label = f"График дежурств — {MONTHS_RU[first_day.month]} {first_day.year}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=hours_col)
    title = ws.cell(row=1, column=1, value=month_label)
    title.fill = _fill(COLORS["header"])
    title.font = _font(bold=True, white=True, size=14)
    title.alignment = _align()
    title.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    ws.row_dimensions[2].height = 36

    h = ws.cell(row=2, column=1, value="Сотрудник")
    h.fill = _fill(COLORS["header"])
    h.font = _font(bold=True, white=True, size=11)
    h.alignment = _align()
    h.border = THIN_BORDER

    hc = ws.cell(row=2, column=2, value="Город")
    hc.fill = _fill(COLORS["header"])
    hc.font = _font(bold=True, white=True, size=9)
    hc.alignment = _align()
    hc.border = THIN_BORDER

    for col_idx, day in enumerate(days, start=3):
        d = day.date
        label = f"{d.day}\n{DAYS_RU[d.weekday()]}"
        cell = ws.cell(row=2, column=col_idx, value=label)
        cell.fill = _fill(COLORS["weekend"] if day.is_holiday else "FFFFFF")
        cell.font = _font(bold=day.is_holiday, size=9)
        cell.alignment = _align()
        cell.border = MONDAY_BORDER if d.weekday() == 0 else THIN_BORDER

    tc = ws.cell(row=2, column=total_col, value="Итого\nдней")
    tc.fill = _fill(COLORS["header"])
    tc.font = _font(bold=True, white=True, size=9)
    tc.alignment = _align()
    tc.border = THIN_BORDER

    hc2 = ws.cell(row=2, column=hours_col, value="Итого\nчасов")
    hc2.fill = _fill(COLORS["header"])
    hc2.font = _font(bold=True, white=True, size=9)
    hc2.alignment = _align()
    hc2.border = THIN_BORDER

    last_row = len(employees) + 2
    dv = DataValidation(type="list", formula1='"Утро,Вечер,Ночь,День,—,Отп"', allow_blank=True)
    dv.sqref = f"C3:{ld}{last_row}"
    ws.add_data_validation(dv)

    for row_idx, emp in enumerate(employees, start=3):
        ws.row_dimensions[row_idx].height = 20
        nc = ws.cell(row=row_idx, column=1, value=_sanitize_cell(emp.name))
        nc.fill = _fill(COLORS["name"])
        nc.font = _font(bold=True, size=10)
        nc.alignment = _align(horizontal="left")
        nc.border = THIN_BORDER

        city_label = "Москва" if emp.city == City.MOSCOW else "Хабаровск"
        city_color = "E8F5E9" if emp.city == City.MOSCOW else "D6E4F0"
        cc = ws.cell(row=row_idx, column=2, value=city_label)
        cc.fill = _fill(city_color)
        cc.font = _font(size=9)
        cc.alignment = _align()
        cc.border = THIN_BORDER

        emp_days = assignments.get(emp.name, {})
        for col_idx, day in enumerate(days, start=3):
            shift_key = emp_days.get(day.date, "day_off")
            label = SHIFT_LABELS.get(shift_key, "?")
            color = CELL_COLORS.get(shift_key, "FFFFFF")
            if day.is_holiday:
                color = _darken(color)
            cell = ws.cell(row=row_idx, column=col_idx, value=label)
            cell.fill = _fill(color)
            cell.font = _font(white=False, size=9)
            cell.alignment = _align()
            cell.border = MONDAY_BORDER if day.date.weekday() == 0 else THIN_BORDER

        dr = f"C{row_idx}:{ld}{row_idx}"
        total_formula = (
            f'=COUNTIF({dr},"Утро")+COUNTIF({dr},"Вечер")+COUNTIF({dr},"Ночь")+COUNTIF({dr},"День")'
        )
        itogo = ws.cell(row=row_idx, column=total_col, value=total_formula)
        itogo.fill = _fill(COLORS["name"])
        itogo.font = _font(bold=True, size=10)
        itogo.alignment = _align()
        itogo.border = THIN_BORDER

        hours_formula = (
            f"=SUMPRODUCT("
            f'(({dr}="Утро")+({dr}="Вечер")+({dr}="Ночь")+({dr}="День")>0)*1'
            f",'{HELPER_SHEET}'!C$2:{ld}$2)"
        )
        hc3 = ws.cell(row=row_idx, column=hours_col, value=hours_formula)
        hc3.fill = _fill(COLORS["name"])
        hc3.font = _font(bold=True, size=10)
        hc3.alignment = _align()
        hc3.border = THIN_BORDER

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    for col_idx in range(3, num_days + 3):
        ws.column_dimensions[get_column_letter(col_idx)].width = 5.5
    ws.column_dimensions[get_column_letter(total_col)].width = 8
    ws.column_dimensions[get_column_letter(hours_col)].width = 8
    ws.freeze_panes = "C3"


def _build_stats_sheet(
    ws: Worksheet,
    stats: list[EmployeeStats],
    schedule: Schedule,
    employees: list[Employee],
    num_days: int,
    ld: str,
) -> None:
    month = schedule.config.month
    year = schedule.config.year
    title = f"Статистика дежурств — {MONTHS_RU[month]} {year}"

    ws.merge_cells("A1:R1")
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = _fill(COLORS["stat_header"])
    title_cell.font = _font(bold=True, white=True, size=14)
    title_cell.alignment = _align()
    title_cell.border = THIN_BORDER
    ws.row_dimensions[1].height = 30

    groups = [
        (1, 1, ""),
        (2, 2, ""),
        (3, 6, "Норма"),
        (7, 10, "Смены"),
        (11, 14, "Отдых"),
        (15, 18, "Нагрузка"),
    ]
    for start, end, label in groups:
        if label:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(row=2, column=start, value=label)
        cell.fill = _fill(COLORS["stat_section"])
        cell.font = _font(bold=True, size=9)
        cell.alignment = _align()
        cell.border = SECTION_BORDER if start in SECTION_COLS else THIN_BORDER
    ws.row_dimensions[2].height = 16

    headers = [
        "Сотрудник",
        "Город",
        "Рабочих\nдней",
        "Норма",
        "±Норма",
        "Часов",
        "Утро",
        "Вечер",
        "Ночь",
        "День",
        "Выходных",
        "Отпуск\n(дней)",
        "Работал в\nвыходные",
        "Работал в\nпраздники",
        "Макс. серия\nработы",
        "Макс. серия\nотдыха",
        "Изол.\nвыходных",
        "Сдвоен.\nвыходных",
    ]
    ws.row_dimensions[3].height = 32
    static_comment = Comment(
        "Пересчитывается при скачивании XLS из UI. "
        "При ручном редактировании XLS в Excel — не обновляется.",
        "System",
    )
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.fill = _fill(COLORS["header"])
        cell.font = _font(bold=True, white=True, size=9)
        cell.alignment = _align()
        cell.border = SECTION_BORDER if col_idx in SECTION_COLS else THIN_BORDER
        if col_idx in (15, 16, 17, 18):
            cell.comment = static_comment

    zero_dash_fmt = '0;-0;"—"'

    for row_idx, st in enumerate(stats, start=4):
        ws.row_dimensions[row_idx].height = 20
        i = row_idx - 4
        sched_row = 3 + i
        helper_row = 6 + i

        nc = ws.cell(row=row_idx, column=1, value=_sanitize_cell(st.name))
        nc.fill = _fill(COLORS["name"])
        nc.font = _font(bold=True, size=10)
        nc.alignment = _align(horizontal="left")
        nc.border = THIN_BORDER

        city_color = "D6E4F0" if st.city == "Хабаровск" else "E8F5E9"
        cc = ws.cell(row=row_idx, column=2, value=st.city)
        cc.fill = _fill(city_color)
        cc.font = _font(size=9)
        cc.alignment = _align()
        cc.border = THIN_BORDER

        dr = f"C{sched_row}:{ld}{sched_row}"

        f_working = f"={_countif_working(SCHED_SHEET, dr)}"
        _stat_cell(ws, row_idx, 3, f_working, COLORS["name"])

        f_norm = f"=ROUND('{HELPER_SHEET}'!$B$5*'{HELPER_SHEET}'!$B${helper_row}/100,0)"
        _stat_cell(ws, row_idx, 4, f_norm, COLORS["name"])

        f_delta = f"=C{row_idx}-D{row_idx}"
        _stat_cell(ws, row_idx, 5, f_delta, COLORS["name"])

        iw = _is_working_array(SCHED_SHEET, dr)
        f_hours = f"=SUMPRODUCT({iw},'{HELPER_SHEET}'!C$2:{ld}$2)"
        _stat_cell(ws, row_idx, 6, f_hours, COLORS["name"])

        f_morning = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"Утро\")"
        _stat_cell(ws, row_idx, 7, f_morning, COLORS["morning"], white=False)
        ws[f"G{row_idx}"].number_format = zero_dash_fmt

        f_evening = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"Вечер\")"
        _stat_cell(ws, row_idx, 8, f_evening, COLORS["evening"], white=True)
        ws[f"H{row_idx}"].number_format = zero_dash_fmt

        f_night = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"Ночь\")"
        _stat_cell(ws, row_idx, 9, f_night, COLORS["night"], white=True)
        ws[f"I{row_idx}"].number_format = zero_dash_fmt

        f_workday = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"День\")"
        _stat_cell(ws, row_idx, 10, f_workday, COLORS["workday"], white=True)
        ws[f"J{row_idx}"].number_format = zero_dash_fmt

        f_dayoff = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"—\")"
        _stat_cell(ws, row_idx, 11, f_dayoff, COLORS["day_off"])

        f_vacation = f"=COUNTIF('{SCHED_SHEET}'!{dr},\"Отп\")"
        _stat_cell(ws, row_idx, 12, f_vacation, COLORS["vacation"], white=True)
        ws[f"L{row_idx}"].number_format = zero_dash_fmt

        f_weekend = f"=SUMPRODUCT('{HELPER_SHEET}'!C$3:{ld}$3*{_is_working_array(SCHED_SHEET, dr)})"
        _stat_cell(ws, row_idx, 13, f_weekend, COLORS["name"])
        ws[f"M{row_idx}"].number_format = zero_dash_fmt

        f_holiday = f"=SUMPRODUCT('{HELPER_SHEET}'!C$4:{ld}$4*{_is_working_array(SCHED_SHEET, dr)})"
        _stat_cell(ws, row_idx, 14, f_holiday, COLORS["name"])
        ws[f"N{row_idx}"].number_format = zero_dash_fmt

        streak_w_color = COLORS["bad"] if st.max_streak_work >= 6 else COLORS["ok"]
        _stat_cell(ws, row_idx, 15, st.max_streak_work, streak_w_color)
        streak_r_color = COLORS["warn"] if st.max_streak_rest >= 3 else COLORS["ok"]
        _stat_cell(ws, row_idx, 16, st.max_streak_rest, streak_r_color)

        if st.isolated_off >= 2:
            iso_color = COLORS["bad"]
        elif st.isolated_off == 1:
            iso_color = COLORS["warn"]
        else:
            iso_color = COLORS["ok"]
        _stat_cell(ws, row_idx, 17, st.isolated_off, iso_color)

        if st.paired_off >= 3:
            paired_color = COLORS["ok"]
        elif st.paired_off >= 1:
            paired_color = COLORS["warn"]
        else:
            paired_color = COLORS["bad"]
        _stat_cell(ws, row_idx, 18, st.paired_off, paired_color)

    total_row = len(stats) + 4
    first_data = 4
    last_data = total_row - 1
    ws.row_dimensions[total_row].height = 22
    _stat_cell(ws, total_row, 1, "ИТОГО по команде", COLORS["total_row"], white=True, bold=True)
    _stat_cell(ws, total_row, 2, "", COLORS["total_row"])

    for col in (3, 6):
        letter = get_column_letter(col)
        _stat_cell(
            ws,
            total_row,
            col,
            f"=SUM({letter}{first_data}:{letter}{last_data})",
            COLORS["total_row"],
            white=True,
            bold=True,
        )

    _stat_cell(ws, total_row, 4, "", COLORS["total_row"])
    _stat_cell(ws, total_row, 5, "", COLORS["total_row"])

    shift_total_styles = {
        7: (COLORS["morning"], False, True),
        8: (COLORS["evening"], True, True),
        9: (COLORS["night"], True, True),
        10: (COLORS["workday"], True, True),
        11: (COLORS["day_off"], False, True),
        12: (COLORS["vacation"], True, True),
        13: (COLORS["total_row"], True, False),
        14: (COLORS["total_row"], True, False),
    }
    for col, (bg, white, bold) in shift_total_styles.items():
        letter = get_column_letter(col)
        _stat_cell(
            ws,
            total_row,
            col,
            f"=SUM({letter}{first_data}:{letter}{last_data})",
            bg,
            white=white,
            bold=bold,
        )

    _stat_cell(ws, total_row, 15, "", COLORS["total_row"])
    _stat_cell(ws, total_row, 16, "", COLORS["total_row"])
    for col in (17, 18):
        letter = get_column_letter(col)
        _stat_cell(
            ws,
            total_row,
            col,
            f"=SUM({letter}{first_data}:{letter}{last_data})",
            COLORS["total_row"],
            white=True,
        )

    delta_range = f"E{first_data}:E{last_data}"
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="equal", formula=["0"], fill=_fill(COLORS["ok"])),
    )
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="between", formula=["-1", "1"], fill=_fill(COLORS["warn"])),
    )
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="greaterThanOrEqual", formula=["2"], fill=_fill(COLORS["bad"])),
    )
    ws.conditional_formatting.add(
        delta_range,
        CellIsRule(operator="lessThanOrEqual", formula=["-2"], fill=_fill(COLORS["bad"])),
    )

    for col_letter in ("M", "N"):
        rng = f"{col_letter}{first_data}:{col_letter}{last_data}"
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="equal", formula=["0"], fill=_fill(COLORS["ok"])),
        )
        ws.conditional_formatting.add(
            rng,
            CellIsRule(operator="greaterThan", formula=["0"], fill=_fill(COLORS["warn"])),
        )

    col_widths = [20, 12, 10, 8, 8, 8, 7, 7, 7, 7, 10, 10, 14, 14, 16, 16, 12, 12]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A3:{get_column_letter(18)}{total_row}"


def _stat_cell(
    ws: Worksheet,
    row: int,
    col: int,
    value: object,
    bg: str,
    white: bool = False,
    bold: bool = False,
) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, white=white, size=10)
    cell.alignment = _align()
    cell.border = SECTION_BORDER if col in SECTION_COLS else THIN_BORDER


def _add_legend(wb: Workbook) -> None:
    ws = wb.create_sheet(title="Легенда")
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 30

    items = [
        ("Утро", "morning", "08:00–17:00 МСК"),
        ("Вечер", "evening", "15:00–00:00 МСК"),
        ("Ночь", "night", "00:00–08:00 МСК (07:00–15:00 ХБ)"),
        ("День", "workday", "Рабочий день 09:00–18:00"),
        ("Отп", "vacation", "Отпуск"),
        ("—", "day_off", "Выходной"),
    ]
    h1 = ws.cell(row=1, column=1, value="Обозн.")
    h1.font = _font(bold=True, size=11)
    h1.border = THIN_BORDER
    h2 = ws.cell(row=1, column=2, value="Описание")
    h2.font = _font(bold=True, size=11)
    h2.border = THIN_BORDER

    for i, (label, key, desc) in enumerate(items, start=2):
        c1 = ws.cell(row=i, column=1, value=label)
        c1.fill = _fill(CELL_COLORS[key])
        c1.font = _font(white=False, bold=True, size=10)
        c1.alignment = _align()
        c1.border = THIN_BORDER
        c2 = ws.cell(row=i, column=2, value=desc)
        c2.font = _font(size=10)
        c2.alignment = _align(horizontal="left")
        c2.border = THIN_BORDER
        ws.row_dimensions[i].height = 20
