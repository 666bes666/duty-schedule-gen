from __future__ import annotations

import tempfile
from datetime import date
from io import BytesIO
from pathlib import Path

import pytest
from openpyxl import Workbook

from duty_schedule.export.xls import export_xls
from duty_schedule.models import (
    City,
    Config,
    DaySchedule,
    Employee,
    Schedule,
    ScheduleType,
    ShiftType,
)
from duty_schedule.xls_import import (
    XlsImportError,
    parse_carry_over_from_xls,
)


def _minimal_employees() -> list[Employee]:
    return [
        Employee(name=f"Москва {i}", city=City.MOSCOW, schedule_type=ScheduleType.FLEXIBLE)
        for i in range(1, 5)
    ] + [
        Employee(name=f"Хабаровск {i}", city=City.KHABAROVSK, schedule_type=ScheduleType.FLEXIBLE)
        for i in range(1, 3)
    ]


def _make_schedule_with_shifts(
    employee_shifts: dict[str, list[str]],
    month: int = 1,
    year: int = 2025,
) -> Schedule:
    employees = _minimal_employees()
    emp_names = {e.name for e in employees}
    for name in employee_shifts:
        if name not in emp_names:
            raise ValueError(f"Unknown employee {name}")

    num_days = len(next(iter(employee_shifts.values())))
    days: list[DaySchedule] = []
    for day_idx in range(num_days):
        d = date(year, month, day_idx + 1)
        day = DaySchedule(date=d)
        for name, shifts in employee_shifts.items():
            shift_key = shifts[day_idx]
            getattr(day, shift_key).append(name)
        days.append(day)

    config = Config(month=month, year=year, seed=42, employees=employees)
    return Schedule(config=config, days=days, metadata={"production_working_days": 21})


def _schedule_to_xls_bytes(schedule: Schedule) -> bytes:
    with tempfile.TemporaryDirectory() as tmpdir:
        path = export_xls(schedule, Path(tmpdir))
        return path.read_bytes()


class TestParseBasic:
    def test_parse_roundtrip(self):
        shifts = {
            "Москва 1": ["morning"] * 5 + ["day_off"] * 2 + ["evening"] * 3,
            "Москва 2": ["evening"] * 5 + ["day_off"] * 2 + ["morning"] * 3,
            "Москва 3": ["morning"] * 3 + ["day_off"] * 2 + ["evening"] * 5,
            "Москва 4": ["evening"] * 3 + ["day_off"] * 2 + ["morning"] * 5,
            "Хабаровск 1": ["night"] * 5 + ["day_off"] * 2 + ["night"] * 3,
            "Хабаровск 2": ["night"] * 3 + ["day_off"] * 2 + ["night"] * 5,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)
        assert len(result) == 6
        names = {co.employee_name for co in result}
        for emp in schedule.config.employees:
            assert emp.name in names


class TestParseEmptyFile:
    def test_empty_bytes(self):
        with pytest.raises(XlsImportError, match="Не удалось открыть"):
            parse_carry_over_from_xls(b"")

    def test_random_bytes(self):
        with pytest.raises(XlsImportError, match="Не удалось открыть"):
            parse_carry_over_from_xls(b"not an xlsx file at all")


class TestUnknownLabels:
    def test_unknown_label_becomes_day_off(self):
        wb = Workbook()
        ws = wb.active
        ws.title = "График дежурств"
        ws.cell(row=1, column=1, value="Заголовок")
        ws.cell(row=2, column=1, value="Сотрудник")
        ws.cell(row=2, column=2, value="Город")
        for d in range(1, 11):
            ws.cell(row=2, column=2 + d, value=str(d))
        ws.cell(row=2, column=13, value="Итого дней")

        ws.cell(row=3, column=1, value="Тест Тестов")
        ws.cell(row=3, column=2, value="Москва")
        for d in range(1, 11):
            ws.cell(row=3, column=2 + d, value="НЕИЗВЕСТНО")

        buf = BytesIO()
        wb.save(buf)

        result = parse_carry_over_from_xls(buf.getvalue())
        assert len(result) == 1
        co = result[0]
        assert co.employee_name == "Тест Тестов"
        assert co.last_shift is None
        assert co.consecutive_off == 10
        assert co.consecutive_working == 0


class TestConsecutiveWorkingAtEnd:
    def test_4_working_days_at_end(self):
        shifts = {
            "Москва 1": ["day_off"] * 6 + ["morning"] * 4,
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["night"] * 10,
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        co = co_map["Москва 1"]
        assert co.consecutive_working == 4
        assert co.consecutive_off == 0
        assert co.last_shift == ShiftType.MORNING


class TestConsecutiveOffAtEnd:
    def test_2_off_at_end(self):
        shifts = {
            "Москва 1": ["morning"] * 8 + ["day_off"] * 2,
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["night"] * 10,
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        co = co_map["Москва 1"]
        assert co.consecutive_off == 2
        assert co.consecutive_working == 0
        assert co.last_shift is None


class TestConsecutiveSameShift:
    def test_3_evenings_at_end(self):
        shifts = {
            "Москва 1": ["morning"] * 5 + ["day_off"] * 2 + ["evening"] * 3,
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["night"] * 10,
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        co = co_map["Москва 1"]
        assert co.last_shift == ShiftType.EVENING
        assert co.consecutive_same_shift == 3
        assert co.consecutive_working == 3


class TestLastShiftDetection:
    def test_detects_morning(self):
        shifts = {
            "Москва 1": ["day_off"] * 9 + ["morning"],
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["night"] * 10,
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        assert co_map["Москва 1"].last_shift == ShiftType.MORNING

    def test_detects_night(self):
        shifts = {
            "Москва 1": ["morning"] * 10,
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["day_off"] * 9 + ["night"],
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        assert co_map["Хабаровск 1"].last_shift == ShiftType.NIGHT

    def test_no_working_days(self):
        shifts = {
            "Москва 1": ["day_off"] * 10,
            "Москва 2": ["morning"] * 10,
            "Москва 3": ["morning"] * 10,
            "Москва 4": ["morning"] * 10,
            "Хабаровск 1": ["night"] * 10,
            "Хабаровск 2": ["night"] * 10,
        }
        schedule = _make_schedule_with_shifts(shifts, month=1, year=2025)
        xls_bytes = _schedule_to_xls_bytes(schedule)
        result = parse_carry_over_from_xls(xls_bytes)

        co_map = {co.employee_name: co for co in result}
        assert co_map["Москва 1"].last_shift is None
        assert co_map["Москва 1"].consecutive_off == 10
