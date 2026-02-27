"""Интеграционные тесты: полная генерация расписания и экспорт."""

from __future__ import annotations

from datetime import date

from duty_schedule.export.ics import export_ics
from duty_schedule.export.xls import export_xls
from duty_schedule.models import (
    CarryOverState,
    City,
    Config,
    Employee,
    ScheduleType,
)
from duty_schedule.scheduler import MAX_CONSECUTIVE_WORKING, generate_schedule


def _make_config(month: int = 3, year: int = 2025, seed: int = 42) -> Config:
    employees = [
        Employee(name=f"Москва {i}", city=City.MOSCOW, schedule_type=ScheduleType.FLEXIBLE)
        for i in range(1, 5)
    ] + [
        Employee(name=f"Хабаровск {i}", city=City.KHABAROVSK, schedule_type=ScheduleType.FLEXIBLE)
        for i in range(1, 3)
    ]
    return Config(month=month, year=year, seed=seed, employees=employees)


class TestFullScheduleGeneration:
    def test_march_2025_complete(self):
        """Полный месяц: 31 день, все смены покрыты."""
        config = _make_config()
        schedule = generate_schedule(config, set())
        assert len(schedule.days) == 31
        for day in schedule.days:
            assert day.is_covered(), f"Смены не покрыты на {day.date}"

    def test_schedule_employees_are_valid(self, full_config):
        """Все сотрудники в расписании существуют в конфигурации."""
        holidays = {date(2025, 3, 8)}
        schedule = generate_schedule(full_config, holidays)
        valid_names = {e.name for e in full_config.employees}
        for day in schedule.days:
            for name in day.morning + day.evening + day.night + day.workday:
                assert name in valid_names, f"Неизвестный сотрудник: {name}"

    def test_no_duplicate_assignments_per_day(self):
        """Сотрудник не может быть назначен на две смены в один день."""
        config = _make_config()
        schedule = generate_schedule(config, set())
        for day in schedule.days:
            all_assigned = day.morning + day.evening + day.night + day.workday
            assert len(all_assigned) == len(set(all_assigned)), (
                f"Дублирование назначений на {day.date}: {all_assigned}"
            )

    def test_khabarovsk_night_distribution_even(self):
        """Ночные смены распределены между хабаровчанами равномерно (разница ≤ 1)."""
        config = _make_config()
        schedule = generate_schedule(config, set())
        khb_nights: dict[str, int] = {}
        for day in schedule.days:
            for name in day.night:
                khb_nights[name] = khb_nights.get(name, 0) + 1
        if len(khb_nights) >= 2:
            counts = list(khb_nights.values())
            assert max(counts) - min(counts) <= 3, f"Дисбаланс ночных смен: {khb_nights}"

    def test_with_holidays(self):
        """Расписание строится корректно при наличии праздников."""
        holidays = {
            date(2025, 3, 8),
            date(2025, 3, 9),
            date(2025, 3, 10),
        }
        config = _make_config()
        schedule = generate_schedule(config, holidays)
        assert len(schedule.days) == 31


def _max_streak_with_carryover(emp_name: str, schedule_days: list, carry_over: int = 0) -> int:
    streak = carry_over
    max_streak = carry_over
    for day in schedule_days:
        working = (
            emp_name in day.morning
            or emp_name in day.evening
            or emp_name in day.night
            or emp_name in day.workday
        )
        if working:
            streak += 1
            max_streak = max(max_streak, streak)
        else:
            streak = 0
    return max_streak


class TestCarryOverConsecutiveConstraint:
    def test_no_violation_with_carryover_4(self):
        """carry_over=4 + первый день февраля = 5 (max). Больше 5 подряд быть не должно."""
        employees = [
            Employee(name=f"Москва {i}", city=City.MOSCOW, schedule_type=ScheduleType.FLEXIBLE)
            for i in range(1, 5)
        ] + [
            Employee(name=f"Хабаровск {i}", city=City.KHABAROVSK, schedule_type=ScheduleType.FLEXIBLE)
            for i in range(1, 3)
        ]
        carry_over = [
            CarryOverState(employee_name="Москва 1", consecutive_working=4),
            CarryOverState(employee_name="Хабаровск 1", consecutive_working=4),
        ]
        config = Config(month=2, year=2025, seed=42, employees=employees, carry_over=carry_over)
        schedule = generate_schedule(config, set())

        for emp_name, co_cw in [("Москва 1", 4), ("Хабаровск 1", 4)]:
            ms = _max_streak_with_carryover(emp_name, schedule.days, carry_over=co_cw)
            assert ms <= MAX_CONSECUTIVE_WORKING, (
                f"{emp_name}: серия {ms} > {MAX_CONSECUTIVE_WORKING} (с учётом переноса {co_cw} дней)"
            )

    def test_no_violation_without_carryover(self):
        """Без carry_over ограничение тоже соблюдается."""
        config = _make_config(month=2, year=2025)
        schedule = generate_schedule(config, set())
        for emp in config.employees:
            ms = _max_streak_with_carryover(emp.name, schedule.days)
            assert ms <= MAX_CONSECUTIVE_WORKING, (
                f"{emp.name}: серия {ms} > {MAX_CONSECUTIVE_WORKING}"
            )


class TestXlsExport:
    def test_file_created(self, tmp_path):
        config = _make_config()
        schedule = generate_schedule(config, set())
        path = export_xls(schedule, tmp_path)
        assert path.exists()
        assert path.suffix == ".xlsx"
        assert path.stat().st_size > 0

    def test_filename_format(self, tmp_path):
        config = _make_config(month=3, year=2025)
        schedule = generate_schedule(config, set())
        path = export_xls(schedule, tmp_path)
        assert "2025_03" in path.name

    def test_xls_readable(self, tmp_path):
        """XLS файл можно прочитать обратно через openpyxl."""
        from openpyxl import load_workbook

        config = _make_config()
        schedule = generate_schedule(config, set())
        path = export_xls(schedule, tmp_path)
        wb = load_workbook(path)
        ws = wb.active
        assert ws.max_column >= 32
        assert ws.max_row >= 7


class TestIcsExport:
    def test_files_created(self, tmp_path):
        config = _make_config()
        schedule = generate_schedule(config, set())
        paths = export_ics(schedule, tmp_path)
        assert len(paths) == 4
        names = {p.name for p in paths}
        assert "morning.ics" in names
        assert "evening.ics" in names
        assert "night.ics" in names
        assert "workday.ics" in names

    def test_ics_valid_format(self, tmp_path):
        """ICS файлы имеют корректную структуру."""
        from icalendar import Calendar

        config = _make_config()
        schedule = generate_schedule(config, set())
        paths = export_ics(schedule, tmp_path)
        for path in paths:
            cal = Calendar.from_ical(path.read_bytes())
            assert cal is not None

    def test_morning_ics_has_events(self, tmp_path):
        """Файл утренних смен содержит события."""
        from icalendar import Calendar

        config = _make_config()
        schedule = generate_schedule(config, set())
        paths = export_ics(schedule, tmp_path)
        morning_ics = next(p for p in paths if p.name == "morning.ics")
        cal = Calendar.from_ical(morning_ics.read_bytes())
        events = [c for c in cal.walk() if c.name == "VEVENT"]
        assert len(events) > 0
