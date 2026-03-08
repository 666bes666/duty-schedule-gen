from __future__ import annotations

import tempfile
from pathlib import Path

from icalendar import Calendar
from openpyxl import load_workbook

from duty_schedule.export.ics import export_ics
from duty_schedule.export.xls import export_xls
from duty_schedule.models import Schedule
from duty_schedule.scheduler import generate_schedule


def _generate(config) -> Schedule:
    holidays = set()
    return generate_schedule(config, holidays)


class TestXlsFormatContract:
    def test_xls_file_created(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            assert path.exists()
            assert path.suffix == ".xlsx"

    def test_xls_has_content(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            assert len(wb.sheetnames) >= 1
            ws = wb.active
            assert ws.max_row > 1
            assert ws.max_column > 1

    def test_schedule_totals_are_formulas(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            ws = wb["График дежурств"]
            num_days = len(schedule.days)
            total_col = num_days + 3
            hours_col = total_col + 1
            for row in range(3, 3 + len(schedule.config.employees)):
                assert str(ws.cell(row=row, column=total_col).value).startswith("=")
                assert str(ws.cell(row=row, column=hours_col).value).startswith("=")

    def test_stats_formulas_reference_schedule(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            ws = wb["Статистика"]
            for row in range(4, 4 + len(schedule.config.employees)):
                for col in range(3, 15):
                    val = str(ws.cell(row=row, column=col).value)
                    assert val.startswith("="), f"Col {col} row {row} not a formula: {val}"

    def test_stats_static_metrics(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            ws = wb["Статистика"]
            for row in range(4, 4 + len(schedule.config.employees)):
                for col in (15, 16, 17, 18):
                    val = ws.cell(row=row, column=col).value
                    assert isinstance(val, int), f"Col {col} row {row} should be int: {val}"

    def test_helper_sheet_hidden(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            assert "_Данные" in wb.sheetnames
            assert wb["_Данные"].sheet_state == "hidden"

    def test_data_validation_exists(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            ws = wb["График дежурств"]
            assert len(ws.data_validations.dataValidation) > 0

    def test_total_row_sum_formulas(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            path = export_xls(schedule, Path(tmpdir))
            wb = load_workbook(path)
            ws = wb["Статистика"]
            total_row = len(schedule.config.employees) + 4
            for col in (3, 6, 7, 8, 9, 10, 11, 12, 13, 14, 17, 18):
                val = str(ws.cell(row=total_row, column=col).value)
                assert "SUM(" in val.upper(), f"Total row col {col}: {val}"


class TestIcsFormatContract:
    def test_ics_files_created(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = export_ics(schedule, Path(tmpdir))
            assert len(paths) == 4
            for p in paths:
                assert p.exists()
                assert p.suffix == ".ics"

    def test_ics_filenames(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = export_ics(schedule, Path(tmpdir))
            names = {p.name for p in paths}
            assert names == {"morning.ics", "evening.ics", "night.ics", "workday.ics"}

    def test_ics_valid_ical(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = export_ics(schedule, Path(tmpdir))
            for p in paths:
                cal = Calendar.from_ical(p.read_bytes())
                assert cal.get("VERSION") == "2.0"
                assert "Duty Schedule" in str(cal.get("PRODID"))

    def test_ics_has_events(self, minimal_config):
        schedule = _generate(minimal_config)
        with tempfile.TemporaryDirectory() as tmpdir:
            paths = export_ics(schedule, Path(tmpdir))
            total_events = 0
            for p in paths:
                cal = Calendar.from_ical(p.read_bytes())
                events = [c for c in cal.walk() if c.name == "VEVENT"]
                total_events += len(events)
            assert total_events > 0
